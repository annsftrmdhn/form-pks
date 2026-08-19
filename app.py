"""
Aplikasi Compile PKS DJBC-TNI AD
=================================
Kanwil upload file Excel PKS mereka -> aplikasi otomatis generate file
Compile PKS sesuai template pusat (3 sheet Laporan, COMPILE DATA per bulan,
CHECKLIST, dan 3 PivotTable native yang otomatis refresh saat file dibuka).

Cara jalan:
    pip install -r requirements.txt
    streamlit run app.py

File template ("template_compile_pks.xlsx") harus ada satu folder dengan
app.py ini.
"""

import io
import re
import zipfile
from copy import copy
from datetime import datetime

import pandas as pd
import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------------------------------------------------------------------
# KONFIGURASI
# ---------------------------------------------------------------------------

TEMPLATE_PATH = "template_compile_pks.xlsx"

# Urutan kanwil resmi (sesuai sheet CHECKLIST di template pusat).
CANONICAL_KANWIL = [
    "Kantor Wilayah DJBC Aceh",
    "Kantor Wilayah DJBC Sumatera Utara",
    "Kantor Wilayah DJBC Riau",
    "Kantor Wilayah DJBC Khusus Kepulauan Riau",
    "Kantor Wilayah DJBC Sumatera Bagian Timur",
    "Kantor Wilayah DJBC Sumatera Bagian Barat",
    "Kantor Wilayah DJBC Banten",
    "Kantor Wilayah DJBC Jakarta",
    "Kantor Wilayah DJBC Jawa Barat",
    "Kantor Wilayah DJBC Jawa Tengah dan DIY",
    "Kantor Wilayah DJBC Jawa Timur I",
    "Kantor Wilayah DJBC Jawa Timur II",
    "Kantor Wilayah DJBC Bali, NTB dan NTT",
    "Kantor Wilayah DJBC Kalimantan Bagian Barat",
    "Kantor Wilayah DJBC Kalimantan Bagian Selatan",
    "Kantor Wilayah DJBC Kalimantan Bagian Timur",
    "Kantor Wilayah DJBC Sulawesi Bagian Selatan",
    "Kantor Wilayah DJBC Sulawesi Bagian Utara",
    "Kantor Wilayah DJBC Maluku",
    "Kantor Wilayah DJBC Khusus Papua",
    "KPU Bea dan Cukai Tipe A Tanjung Priok",
    "KPU Bea dan Cukai Tipe B Batam",
    "KPU Bea dan Cukai Tipe C Soekarno-Hatta",
]

SHEETS = {
    "LAPORAN SOSIALISASI": {"header_row": 3, "last_col": 12, "text_col": 3},
    "LAPORAN PUBLIKASI": {"header_row": 3, "last_col": 10, "text_col": 3},
    "LAPORAN SIARAN PERS": {"header_row": 3, "last_col": 9, "text_col": 3},
}
KANWIL_COL = 4   # kolom "Kanwil DJBC" (D) di semua sheet laporan
SATKER_COL = 5   # kolom "Satuan DJBC" (E) -> dipakai untuk rekap per satker

CHECKLIST_SHEET = "CHECKLIST"

REKAP_LABELS = {
    "LAPORAN SOSIALISASI": "Sosialisasi",
    "LAPORAN PUBLIKASI": "Publikasi",
    "LAPORAN SIARAN PERS": "Siaran Pers",
}


# ---------------------------------------------------------------------------
# 1. BACA FILE KANWIL
# ---------------------------------------------------------------------------

def find_header_row(ws, search_rows=15) -> int:
    """Cari baris header secara otomatis: baris yang kolom B-nya persis 'No'."""
    for r in range(1, search_rows + 1):
        val = ws.cell(row=r, column=2).value
        if val is not None and str(val).strip().lower() == "no":
            return r
    return 3


def is_valid_tni_ad(unit_name) -> bool:
    """Cek apakah instansi merupakan satuan TNI AD yang valid. Jika instansi lain (Polri, Pemda, dll), kembalikan False."""
    if unit_name is None:
        return True
        
    name = str(unit_name).lower().strip()
    if name in ("", "-", "nihil"):
        return True
        
    blacklist = [
        "satpol", "pol pp", "polisi", "polres", "polsek", "polda", "polri", "brimob", "sabhara", "bhabinkamtibmas",
        "tni al", "lanal", "lantamal", "koarmada", "marinir", "angkatan laut",
        "tni au", "lanud", "koopsau", "paskhas", "kopasgat", "angkatan udara",
        "dishub", "kejaksaan", "bnn", "karantina", "imigrasi", "pemda", "satlinmas", "linmas", "polhut", "bakamla", "kplp"
    ]
    
    whitelist = [
        "kodam", "korem", "kodim", "koramil", "babinsa", "danrem", "dandim", "danramil", "pangdam",
        "yonif", "yonkav", "yonarmed", "yonarhanud", "yonzipur", "kikav", "denzipur", "batalyon",
        "denpom", "pomdam", "puspom", "pom", 
        "kopassus", "kostrad", "tni ad", "tni-ad", "ad", "angkatan darat", "mabesad",
        "pamtas", "satgas", "brigif", "grup", "kipan", "pos",
        "rindam", "secata", "secaba", "akmil", "seskoad",
        "pussen", "pusdik", "bais", "tni"
    ]
    
    has_black = any(b in name for b in blacklist)
    has_white = any(w in name for w in whitelist)
    
    if has_black and not has_white:
        return False
        
    return True



def read_kanwil_file(file_bytes: bytes):
    """Baca 1 file excel kanwil -> {sheet_name: [row_values...]} + nama kanwil terdeteksi."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    extracted = {}
    detected_names = []

    for sheet_name, cfg in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            extracted[sheet_name] = []
            continue
        ws = wb[sheet_name]
        hr = find_header_row(ws)
        rows = []
        for r in range(hr + 1, ws.max_row + 1):
            values = [ws.cell(row=r, column=c).value for c in range(2, cfg["last_col"] + 1)]
            
            # Deteksi kanwil name terlebih dahulu
            kw = ws.cell(row=r, column=KANWIL_COL).value
            if kw:
                detected_names.append(str(kw).strip())

            if all(v in (None, "") for v in values):
                continue
                

            # Filter baris yang Satuan TNI AD-nya berisi instansi selain TNI AD
            tni_ad_val = values[4] if len(values) > 4 else None
            if not is_valid_tni_ad(tni_ad_val):
                continue
                
            rows.append(values)
        extracted[sheet_name] = rows

    kanwil_guess = None
    if detected_names:
        kanwil_guess = max(set(detected_names), key=detected_names.count)
    return extracted, kanwil_guess


def match_canonical(name: str):
    """Cocokkan nama kanwil hasil deteksi ke daftar resmi (toleran spasi/typo kecil)."""
    if not name:
        return None
    norm = re.sub(r"\s+", " ", name).strip().lower()
    for canon in CANONICAL_KANWIL:
        if re.sub(r"\s+", " ", canon).strip().lower() == norm:
            return canon
    for canon in CANONICAL_KANWIL:
        if norm in canon.lower() or canon.lower() in norm:
            return canon
    return None


# ---------------------------------------------------------------------------
# 2. GABUNGKAN KE TEMPLATE
# ---------------------------------------------------------------------------

def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)


def build_compile_workbook(template_bytes: bytes, data_by_kanwil: dict, period_label: str):
    """data_by_kanwil: {canonical_kanwil_name: {sheet_name: [rows]}}"""
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    sheet_last_row = {}

    # Cari nama sheet COMPILE DATA (misal COMPILE DATA JUNI)
    compile_sheet_name = None
    for s in wb.sheetnames:
        if "COMPILE DATA" in s.upper():
            compile_sheet_name = s
            break
    if not compile_sheet_name:
        compile_sheet_name = "COMPILE DATA"
        wb.create_sheet(compile_sheet_name)

    # Rename sheet COMPILE DATA ke nama periode terbaru
    new_compile_name = f"COMPILE DATA {period_label.upper()}".strip()
    if compile_sheet_name in wb.sheetnames:
        wb[compile_sheet_name].title = new_compile_name

    # Write data ke 3 sheet laporan
    for sheet_name, cfg in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # Update Judul di Row 1
        title_val = f"Laporan {REKAP_LABELS[sheet_name]} Bulan {period_label} (PKS DJBC-TNI AD)"
        for col_idx in range(1, 10):
            if ws.cell(row=1, column=col_idx).value:
                ws.cell(row=1, column=col_idx, value=title_val)
                break

        hr = cfg["header_row"]
        style_row = hr + 1
        last_col = cfg["last_col"]

        # Hapus sisa baris lama jika ada
        if ws.max_row > hr:
            ws.delete_rows(hr + 1, ws.max_row - hr)

        out_row = hr + 1
        no_counter = 0

        for kanwil in CANONICAL_KANWIL:
            rows = data_by_kanwil.get(kanwil, {}).get(sheet_name, [])
            if not rows:
                continue
            no_counter += 1
            kanwil_start_row = out_row
            for i, values in enumerate(rows):
                target_row = out_row
                for c in range(2, last_col + 1):
                    cell = ws.cell(row=target_row, column=c)
                    val = values[c - 2] if (c - 2) < len(values) else None
                    
                    # STANDARDIZE KANWIL FOR PIVOT TABLE
                    if c == KANWIL_COL:
                        val = kanwil
                            
                    cell.value = val
                # Kolom "No" (B) hanya diisi di baris pertama tiap kanwil
                ws.cell(row=target_row, column=2, value=no_counter if i == 0 else None)
                out_row += 1

            kanwil_end_row = out_row - 1
            # Merge sel dalam satu Kanwil jika lebih dari 1 baris
            if kanwil_start_row < kanwil_end_row:
                # Kolom No (B) dan Kanwil DJBC (D) selalu di-merge
                ws.merge_cells(start_row=kanwil_start_row, start_column=2,
                               end_row=kanwil_end_row, end_column=2)
                ws.cell(row=kanwil_start_row, column=2).alignment = Alignment(
                    horizontal="center", vertical="center")
                ws.merge_cells(start_row=kanwil_start_row, start_column=KANWIL_COL,
                               end_row=kanwil_end_row, end_column=KANWIL_COL)
                ws.cell(row=kanwil_start_row, column=KANWIL_COL).alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True)

                # Kolom lain: merge jika semua baris punya nilai yang sama
                for c in range(2, last_col + 1):
                    if c == 2 or c == KANWIL_COL:
                        continue  # sudah di-merge di atas
                    vals_in_block = []
                    for r in range(kanwil_start_row, kanwil_end_row + 1):
                        v = ws.cell(row=r, column=c).value
                        vals_in_block.append(v)
                    # Merge jika semua nilai sama dan tidak None/kosong
                    if (len(set(str(v) if v is not None else "" for v in vals_in_block)) == 1
                            and vals_in_block[0] is not None
                            and str(vals_in_block[0]).strip() != ""):
                        ws.merge_cells(start_row=kanwil_start_row, start_column=c,
                                       end_row=kanwil_end_row, end_column=c)
                        ws.cell(row=kanwil_start_row, column=c).alignment = Alignment(
                            horizontal="center", vertical="center", wrap_text=True)

        sheet_last_row[sheet_name] = max(out_row - 1, hr + 1)

    # ---- COMPILE DATA: susun rekap 5 bagian lengkap dengan RUMUS EXCEL NATIVE ----
    compile_ws = wb[new_compile_name]
    write_compile_summary(compile_ws, data_by_kanwil, period_label)

    # ---- CHECKLIST: tandai status kanwil ----
    if CHECKLIST_SHEET in wb.sheetnames:
        checklist_ws = wb[CHECKLIST_SHEET]
        mark_checklist(checklist_ws, data_by_kanwil)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue(), sheet_last_row


def write_compile_summary(ws, data_by_kanwil, period_label: str):
    """Generates the full 5-section COMPILE DATA sheet matching the original compile Excel file."""

    # Clear worksheet content completely
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row + 10)
        
    # Rapihkan lebar kolom
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25


    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11, bold=False)
    title_font = Font(name="Calibri", size=14, bold=True)

    header_fill = PatternFill("solid", fgColor="D8D8D8")
    rekap_fill = PatternFill("solid", fgColor="BDD7EE")
    yellow_fill = PatternFill("solid", fgColor="FFFF00")

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")

    # Row 1 Title (di col E, sesuai template referensi)
    title_cell = ws.cell(row=1, column=5, value=f"Data Total Kegiatan Komunikasi Unit Vertikal PKS DJBC-TNI AD Bulan {period_label}")
    title_cell.font = title_font
    title_cell.alignment = center_align

    sections = [
        ("LAPORAN SOSIALISASI", "Sosialisasi"),
        ("LAPORAN PUBLIKASI", "Publikasi"),
        ("LAPORAN SIARAN PERS", "Siaran Pers"),
    ]

    current_row = 3
    tot_cells = {}
    kanwil_cell_map = {label: {} for _, label in sections}
    satker_cell_map = {label: {} for _, label in sections}

    for sheet_name, label in sections:
        cfg = SHEETS[sheet_name]
        text_idx = cfg["text_col"] - 2
        satker_idx = SATKER_COL - 2

        # Collect non-NIHIL counts per (kanwil, satker)
        sec_data = {}
        for kanwil in CANONICAL_KANWIL:
            rows = data_by_kanwil.get(kanwil, {}).get(sheet_name, [])
            for vals in rows:
                if all(v is None or str(v).strip() == "" for v in vals):
                    continue
                tval = vals[text_idx] if text_idx < len(vals) else None
                satker = vals[satker_idx] if satker_idx < len(vals) else None

                if tval is not None and str(tval).strip().upper() == "NIHIL":
                    continue

                sname = str(satker).strip() if (satker is not None and str(satker).strip() != "") else "(Tanpa Satker)"
                if kanwil not in sec_data:
                    sec_data[kanwil] = {}
                sec_data[kanwil][sname] = sec_data[kanwil].get(sname, 0) + 1

        r1 = current_row
        r2 = current_row + 1

        # Table Header
        ws.merge_cells(start_row=r1, start_column=1, end_row=r2, end_column=1)
        ws.cell(row=r1, column=1, value="No.").font = bold_font
        ws.cell(row=r1, column=1).alignment = center_align

        ws.merge_cells(start_row=r1, start_column=2, end_row=r1, end_column=5)
        ws.cell(row=r1, column=2, value=label).font = bold_font
        ws.cell(row=r1, column=2).alignment = center_align

        sub_headers = ["Kanwil", "Satker", "Jumlah per Satker", "Total per Kanwil"]
        for col_i, sh in enumerate(sub_headers, start=2):
            cell = ws.cell(row=r2, column=col_i, value=sh)
            cell.font = bold_font
            cell.alignment = center_align

        for row_idx in range(r1, r2 + 1):
            for col_idx in range(1, 6):
                c_cell = ws.cell(row=row_idx, column=col_idx)
                c_cell.border = border
                c_cell.fill = header_fill

        current_row = r2 + 1
        first_sec_row = current_row
        no_counter = 0

        for kanwil in CANONICAL_KANWIL:
            if kanwil not in sec_data:
                continue
            no_counter += 1
            satker_dict = sec_data[kanwil]
            r_start = current_row

            for i, (satker, count) in enumerate(satker_dict.items()):
                r = current_row
                ws.cell(row=r, column=1, value=no_counter if i == 0 else None).alignment = center_align
                ws.cell(row=r, column=2, value=kanwil if i == 0 else None).alignment = left_align
                ws.cell(row=r, column=3, value=satker).alignment = left_align
                ws.cell(row=r, column=4, value=count).alignment = center_align

                for c in range(1, 6):
                    c_cell = ws.cell(row=r, column=c)
                    c_cell.border = border
                    c_cell.font = regular_font

                satker_cell_map[label][(kanwil, satker)] = (r, 4)
                current_row += 1

            r_end = current_row - 1
            top_cell = ws.cell(row=r_start, column=5)
            top_cell.alignment = center_align
            top_cell.font = regular_font
            if r_start == r_end:
                top_cell.value = f"=D{r_start}"
            else:
                top_cell.value = f"=SUM(D{r_start}:D{r_end})"
            
            # Merge sel untuk No, Kanwil, dan Total per Kanwil
            if r_start < r_end:
                ws.merge_cells(start_row=r_start, start_column=1, end_row=r_end, end_column=1)
                ws.merge_cells(start_row=r_start, start_column=2, end_row=r_end, end_column=2)
                ws.merge_cells(start_row=r_start, start_column=5, end_row=r_end, end_column=5)

            kanwil_cell_map[label][kanwil] = (r_start, 5)

        last_sec_row = current_row - 1
        if last_sec_row < first_sec_row:
            last_sec_row = first_sec_row

        # Total Row
        r_tot = current_row
        ws.merge_cells(start_row=r_tot, start_column=1, end_row=r_tot, end_column=4)
        t_cell = ws.cell(row=r_tot, column=1, value=f"TOTAL KEGIATAN {label.upper()} ")
        t_cell.font = bold_font
        t_cell.alignment = left_align

        tot_val_cell = ws.cell(row=r_tot, column=5, value=f"=SUM(E{first_sec_row}:E{last_sec_row})")
        tot_val_cell.font = bold_font
        tot_val_cell.alignment = center_align
        tot_cells[label] = f"E{r_tot}"

        for c in range(1, 6):
            c_cell = ws.cell(row=r_tot, column=c)
            c_cell.fill = header_fill
            c_cell.border = border

        current_row += 2  # skip 1 row

    # Grand Total Row
    r_gt = current_row - 1
    ws.merge_cells(start_row=r_gt, start_column=1, end_row=r_gt, end_column=4)
    gt_label = ws.cell(row=r_gt, column=1, value="TOTAL KEGIATAN PKS DJBC-TNI AD ")
    gt_label.font = bold_font
    gt_label.alignment = left_align

    gt_formula = "=SUM(" + tot_cells["Sosialisasi"] + "," + tot_cells["Publikasi"] + "," + tot_cells["Siaran Pers"] + ")"
    gt_val = ws.cell(row=r_gt, column=5, value=gt_formula)
    gt_val.font = bold_font
    gt_val.alignment = center_align

    for c in range(1, 6):
        c_cell = ws.cell(row=r_gt, column=c)
        c_cell.fill = yellow_fill
        c_cell.border = border

    current_row += 3  # skip 2 rows

    # SECTION 4: Rekap Seluruh Kegiatan Per Kanwil
    active_kanwils = []
    for k in CANONICAL_KANWIL:
        if any(k in kanwil_cell_map[l] for l in ["Sosialisasi", "Publikasi", "Siaran Pers"]):
            active_kanwils.append(k)

    r1 = current_row
    r2 = current_row + 1
    ws.merge_cells(start_row=r1, start_column=1, end_row=r2, end_column=1)
    ws.cell(row=r1, column=1, value="No.").font = bold_font
    ws.cell(row=r1, column=1).alignment = center_align

    ws.merge_cells(start_row=r1, start_column=2, end_row=r1, end_column=3)
    ws.cell(row=r1, column=2, value="Rekap Seluruh Kegiatan Per Kanwil").font = bold_font
    ws.cell(row=r1, column=2).alignment = center_align

    ws.cell(row=r2, column=2, value="Kanwil").font = bold_font
    ws.cell(row=r2, column=2).alignment = center_align

    ws.cell(row=r2, column=3, value="Total Seluruh Kegiatan").font = bold_font
    ws.cell(row=r2, column=3).alignment = center_align

    for row_idx in range(r1, r2 + 1):
        for col_idx in range(1, 4):
            c_cell = ws.cell(row=row_idx, column=col_idx)
            c_cell.border = border
            c_cell.fill = rekap_fill

    current_row = r2 + 1
    first_k_rekap = current_row

    for no_idx, kanwil in enumerate(active_kanwils, 1):
        r = current_row
        ws.cell(row=r, column=1, value=no_idx).alignment = center_align
        ws.cell(row=r, column=2, value=kanwil).alignment = left_align

        refs = []
        for l in ["Sosialisasi", "Publikasi", "Siaran Pers"]:
            if kanwil in kanwil_cell_map[l]:
                kr, kc = kanwil_cell_map[l][kanwil]
                refs.append(f"E{kr}")
        ws.cell(row=r, column=3, value="=" + "+".join(refs)).alignment = center_align

        for c in range(1, 4):
            ws.cell(row=r, column=c).border = border
            ws.cell(row=r, column=c).font = regular_font
        current_row += 1

    last_k_rekap = max(current_row - 1, first_k_rekap)
    r_k_tot = current_row
    ws.merge_cells(start_row=r_k_tot, start_column=1, end_row=r_k_tot, end_column=2)
    ws.cell(row=r_k_tot, column=1, value=f"TOTAL KEGIATAN BULAN {period_label.upper()}").font = bold_font
    ws.cell(row=r_k_tot, column=1).alignment = center_align

    tot_k_cell = ws.cell(row=r_k_tot, column=3, value=f"=SUM(C{first_k_rekap}:C{last_k_rekap})")
    tot_k_cell.font = bold_font
    tot_k_cell.alignment = center_align

    avg_k_cell = ws.cell(row=r_k_tot, column=4, value=f"=AVERAGE(C{first_k_rekap}:C{last_k_rekap})")
    avg_k_cell.font = regular_font
    avg_k_cell.alignment = center_align

    for c in range(1, 4):
        ws.cell(row=r_k_tot, column=c).border = border

    current_row += 3  # skip 2 rows

    # SECTION 5: Rekap Seluruh Kegiatan Per Kanwil & Satker
    r1 = current_row
    r2 = current_row + 1
    ws.merge_cells(start_row=r1, start_column=1, end_row=r2, end_column=1)
    ws.cell(row=r1, column=1, value="No.").font = bold_font
    ws.cell(row=r1, column=1).alignment = center_align

    ws.merge_cells(start_row=r1, start_column=2, end_row=r1, end_column=3)
    ws.cell(row=r1, column=2, value="Rekap Seluruh Kegiatan Per Kanwil").font = bold_font
    ws.cell(row=r1, column=2).alignment = center_align

    ws.cell(row=r2, column=2, value="Kanwil").font = bold_font
    ws.cell(row=r2, column=2).alignment = center_align

    ws.cell(row=r2, column=3, value="Total Seluruh Kegiatan").font = bold_font
    ws.cell(row=r2, column=3).alignment = center_align

    for row_idx in range(r1, r2 + 1):
        for col_idx in range(1, 4):
            ws.cell(row=row_idx, column=col_idx).border = border

    current_row = r2 + 1
    kanwil_head_rows = []

    for no_idx, kanwil in enumerate(active_kanwils, 1):
        r_head = current_row
        kanwil_head_rows.append(r_head)
        ws.cell(row=r_head, column=1, value=no_idx).alignment = center_align
        ws.cell(row=r_head, column=2, value=kanwil).alignment = left_align

        for c in range(1, 4):
            ws.cell(row=r_head, column=c).border = border
            ws.cell(row=r_head, column=c).font = bold_font

        current_row += 1
        satker_order = []
        for l in ["Sosialisasi", "Publikasi", "Siaran Pers"]:
            for (k, s) in satker_cell_map[l].keys():
                if k == kanwil and s not in satker_order:
                    satker_order.append(s)

        satker_start = current_row
        for sname in satker_order:
            r_sat = current_row
            ws.cell(row=r_sat, column=2, value=sname).alignment = left_align

            s_refs = []
            for l in ["Sosialisasi", "Publikasi", "Siaran Pers"]:
                if (kanwil, sname) in satker_cell_map[l]:
                    sr, sc = satker_cell_map[l][(kanwil, sname)]
                    s_refs.append(f"D{sr}")

            ws.cell(row=r_sat, column=3, value="=" + "+".join(s_refs)).alignment = center_align

            for c in range(1, 4):
                ws.cell(row=r_sat, column=c).border = border
                ws.cell(row=r_sat, column=c).font = regular_font
            current_row += 1

        satker_end = current_row - 1
        head_c_cell = ws.cell(row=r_head, column=3)
        head_c_cell.alignment = center_align
        if satker_start <= satker_end:
            if satker_start == satker_end:
                head_c_cell.value = f"=C{satker_start}"
            else:
                head_c_cell.value = f"=SUM(C{satker_start}:C{satker_end})"

            # Merge col A untuk blok kanwil+satker (sesuai template referensi)
            ws.merge_cells(start_row=r_head, start_column=1, end_row=satker_end, end_column=1)

    r_sat_tot = current_row
    ws.merge_cells(start_row=r_sat_tot, start_column=1, end_row=r_sat_tot, end_column=2)
    ws.cell(row=r_sat_tot, column=1, value=f"TOTAL KEGIATAN BULAN {period_label.upper()}").font = bold_font
    ws.cell(row=r_sat_tot, column=1).alignment = center_align

    if kanwil_head_rows:
        sum_h_refs = [f"C{h}" for h in kanwil_head_rows]
        tot_s_cell = ws.cell(row=r_sat_tot, column=3, value="=" + "+".join(sum_h_refs))
    else:
        tot_s_cell = ws.cell(row=r_sat_tot, column=3, value=0)
    tot_s_cell.font = bold_font
    tot_s_cell.alignment = center_align

    for c in range(1, 4):
        ws.cell(row=r_sat_tot, column=c).border = border


def mark_checklist(ws, data_by_kanwil):
    for r in range(4, ws.max_row + 1):
        kanwil = ws.cell(row=r, column=3).value
        if not kanwil:
            continue
        canon = match_canonical(str(kanwil))
        sudah = canon in data_by_kanwil if canon else False
        ws.cell(row=r, column=4, value="Sudah" if sudah else "-")
        ws.cell(row=r, column=5, value="Sudah" if sudah else "-")
        ws.cell(row=r, column=6, value="Sudah" if sudah else "Belum")


# ---------------------------------------------------------------------------
# 3. PATCH RANGE PIVOTTABLE (biar tinggal "Refresh All" di Excel)
# ---------------------------------------------------------------------------

COL_LETTERS = {
    "LAPORAN SOSIALISASI": "L",
    "LAPORAN PUBLIKASI": "J",
    "LAPORAN SIARAN PERS": "I",
}


def patch_pivot_ranges(xlsx_bytes: bytes, sheet_last_row: dict) -> bytes:
    """Update <worksheetSource ref="B3:Lxxx"/> di tiap pivotCacheDefinition
    supaya mencakup baris data terbaru, dan set refreshOnLoad="1" supaya
    Excel otomatis refresh saat file dibuka."""
    buf_in = io.BytesIO(xlsx_bytes)
    zin = zipfile.ZipFile(buf_in, "r")
    buf_out = io.BytesIO()
    zout = zipfile.ZipFile(buf_out, "w", zipfile.ZIP_DEFLATED)

    for item in zin.infolist():
        data = zin.read(item.filename)
        if re.match(r"xl/pivotCache/pivotCacheDefinition\d+\.xml", item.filename):
            text = data.decode("utf-8")

            def repl_ref(m):
                sheet = m.group("sheet")
                last_row = sheet_last_row.get(sheet)
                col = COL_LETTERS.get(sheet, "L")
                if last_row:
                    return f'<worksheetSource ref="B3:{col}{last_row}" sheet="{sheet}"/>'
                return m.group(0)

            text = re.sub(
                r'<worksheetSource\s+ref="[^"]+"\s+sheet="(?P<sheet>[^"]+)"\s*/>',
                repl_ref,
                text,
            )
            if 'refreshOnLoad="1"' not in text:
                text = text.replace("<pivotCacheDefinition ", '<pivotCacheDefinition refreshOnLoad="1" ', 1)
            data = text.encode("utf-8")
        zout.writestr(item, data)

    zin.close()
    zout.close()
    return buf_out.getvalue()


# ---------------------------------------------------------------------------
# STREAMLIT UI
# ---------------------------------------------------------------------------

st.set_page_config(page_title="Generate Compile PKS", page_icon="📊", layout="centered")

# --- AUTHENTICATION ---
if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False

def login_form():
    st.title("🔒 Login PKS")
    st.write("Silakan login untuk mengakses aplikasi Compile PKS.")
    with st.form("login_form"):
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Login")
        if submitted:
            # Hardcoded credentials
            if username == "admin" and password == "pksdjbc2026":
                st.session_state["logged_in"] = True
                st.rerun()
            else:
                st.error("Username atau password salah")

if not st.session_state["logged_in"]:
    login_form()
    st.stop()

# --- SIDEBAR / LOGOUT ---
with st.sidebar:
    st.write("Logged in as: **admin**")
    if st.button("Sign Out"):
        st.session_state["logged_in"] = False
        st.rerun()

# --- MAIN APP ---
st.title("📊 Generate Compile PKS DJBC-TNI AD")
st.caption(
    "Upload file Excel PKS dari masing-masing Kanwil, aplikasi akan otomatis "
    "menggabungkannya sesuai format Compile PKS pusat (termasuk PivotTable & rumus native Excel)."
)

period_label = st.text_input("Label periode (untuk nama file output)", value="Juli 2026")

uploaded_files = st.file_uploader(
    "Upload file Excel PKS dari Kanwil (bisa lebih dari satu sekaligus)",
    type=["xlsx"],
    accept_multiple_files=True,
)

if uploaded_files:
    st.subheader("Konfirmasi nama Kanwil per file")
    st.caption("Aplikasi mencoba mendeteksi otomatis dari isi file — cek dan sesuaikan kalau perlu.")

    data_by_kanwil = {}

    for f in uploaded_files:
        file_bytes = f.getvalue()
        extracted, guess = read_kanwil_file(file_bytes)
        canon_guess = match_canonical(guess) if guess else None
        default_idx = CANONICAL_KANWIL.index(canon_guess) if canon_guess in CANONICAL_KANWIL else 0

        selected = st.selectbox(
            f"📄 {f.name}",
            options=CANONICAL_KANWIL,
            index=default_idx,
            key=f"map_{f.name}",
        )
        if selected in data_by_kanwil:
            st.warning(f"⚠️ '{selected}' sudah dipilih untuk file lain — akan ditimpa oleh file ini.")
        data_by_kanwil[selected] = extracted

    st.divider()

    if st.button("🚀 Generate Compile PKS", type="primary"):
        try:
            with open(TEMPLATE_PATH, "rb") as tf:
                template_bytes = tf.read()
        except FileNotFoundError:
            st.error(
                f"File template '{TEMPLATE_PATH}' tidak ditemukan. "
                "Pastikan file itu ada di folder yang sama dengan app.py."
            )
            st.stop()

        with st.spinner("Menggabungkan data & menyusun Compile PKS..."):
            xlsx_bytes, sheet_last_row = build_compile_workbook(
                template_bytes, data_by_kanwil, period_label
            )
            xlsx_bytes = patch_pivot_ranges(xlsx_bytes, sheet_last_row)

        st.success(
            f"Selesai! {len(data_by_kanwil)} Kanwil berhasil digabungkan. "
            "File yang di-download sudah berisi rumus Excel native dan PivotTable terkonfigurasi."
        )

        fname = f"COMPILE PKS DJBC TNI AD {period_label}.xlsx".replace("  ", " ")
        st.download_button(
            "⬇️ Download Compile PKS",
            data=xlsx_bytes,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
else:
    st.info("Silakan upload minimal satu file Excel PKS dari Kanwil untuk mulai.")
